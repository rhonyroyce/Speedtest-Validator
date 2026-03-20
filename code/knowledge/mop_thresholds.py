"""MOP threshold loader — runtime loading of DAS_Validation_Thresholds.xlsx.

CRITICAL: Speed test thresholds are NEVER hardcoded. This module loads them
from the Excel file at runtime, providing the 5-dimensional lookup that
threshold_engine.py uses for pass/fail determination.

Reference: DAS_Validation_Thresholds.xlsx (SISO + MIMO sheets, 16 BW combos each)
Implementation: Claude Code Prompt 5 (Knowledge Engine)
"""
from pathlib import Path

import openpyxl


# Sheet names in DAS_Validation_Thresholds.xlsx
SHEET_NAMES = {
    "service_mode": "Service Mode Thresholds",
    "siso": "SISO Speed Test",
    "mimo": "MIMO Speed Test",
    "lookup_guide": "Lookup Guide",
    "site_config": "SFY0803A Config",
}

# Column mapping for SISO/MIMO sheets
# Row structure: Config | BW LTE | BW NR C1 | BW NR C2 | LTE DL Min | LTE DL Max |
#                NR DL Min | NR DL Max | EN-DC DL Min | EN-DC DL Max | NR-DC DL Min | NR-DC DL Max |
#                LTE UL Min | LTE UL Max | NR UL Min | NR UL Max | EN-DC UL Min | EN-DC UL Max |
#                UL SINR15 | UL SINR19 | UL SINR20 | UL SINR22 | UL SINR24 | UL SINR26 | UL SINR28 | UL SINR30
COLUMN_MAP = {
    "config": 0,      # "SISO" or "MIMO"
    "bw_lte": 1,      # MHz
    "bw_nr_c1": 2,    # MHz
    "bw_nr_c2": 3,    # MHz
    # DL thresholds (Min, Max pairs)
    "lte_dl_min": 4,
    "lte_dl_max": 5,
    "nr_dl_min": 6,
    "nr_dl_max": 7,
    "endc_dl_min": 8,
    "endc_dl_max": 9,
    "nrdc_dl_min": 10,
    "nrdc_dl_max": 11,
    # UL thresholds (Min, Max pairs)
    "lte_ul_min": 12,
    "lte_ul_max": 13,
    "nr_ul_min": 14,
    "nr_ul_max": 15,
    "endc_ul_min": 16,
    "endc_ul_max": 17,
    # Progressive UL by SINR (EN-DC specific)
    "ul_sinr_15": 18,
    "ul_sinr_19": 19,
    "ul_sinr_20": 20,
    "ul_sinr_22": 21,
    "ul_sinr_24": 22,
    "ul_sinr_26": 23,
    "ul_sinr_28": 24,
    "ul_sinr_30": 25,
}

# Progressive UL SINR levels
PROGRESSIVE_UL_SINR_LEVELS = [15, 19, 20, 22, 24, 26, 28, 30]


def load_threshold_excel(excel_path: str | Path) -> dict:
    """Load all threshold data from DAS_Validation_Thresholds.xlsx.

    Args:
        excel_path: Path to the threshold Excel file

    Returns:
        Dict with keys:
            'service_mode': {param: {min, max, unit}}
            'siso': [list of row dicts with all columns]
            'mimo': [list of row dicts with all columns]

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheets are missing or data format unexpected
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Threshold Excel not found: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    # --- Service Mode Thresholds ---
    sm_sheet_name = SHEET_NAMES["service_mode"]
    if sm_sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing sheet: {sm_sheet_name}")
    ws_sm = wb[sm_sheet_name]
    sm_rows = list(ws_sm.iter_rows(values_only=True))
    # Row 0 = header, Row 1 = Radio End, Row 2 = Antenna End
    service_mode = {}
    for row in sm_rows[1:]:
        location = row[0]
        service_mode[location] = {
            "rsrp_min": row[1],
            "rsrp_max": row[2],
            "sinr_min": row[3],
            "rsrq_min": row[4],
            "rsrq_max": row[5],
            "tx_power": row[6],
        }

    # --- Helper to parse speed test sheet ---
    def _parse_speed_sheet(sheet_name: str) -> list[dict]:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Row 0 = header, rows 1-16 = data
        data_rows = []
        for row in rows[1:]:
            if row[0] is None:
                continue
            row_dict = {}
            for col_name, col_idx in COLUMN_MAP.items():
                if col_idx < len(row):
                    row_dict[col_name] = row[col_idx]
                else:
                    row_dict[col_name] = None
            data_rows.append(row_dict)
        return data_rows

    siso_rows = _parse_speed_sheet(SHEET_NAMES["siso"])
    mimo_rows = _parse_speed_sheet(SHEET_NAMES["mimo"])

    # --- Physical Thresholds Lookup (optional) ---
    physical = []
    phys_sheet = "Physical Thresholds Lookup"
    if phys_sheet in wb.sheetnames:
        ws_phys = wb[phys_sheet]
        phys_rows = list(ws_phys.iter_rows(values_only=True))
        for row in phys_rows[1:]:
            if row[0] is None:
                continue
            physical.append({
                "parameter": row[0],
                "band": row[1],
                "condition": row[2],
                "min_value": row[3],
                "max_value": row[4],
                "unit": row[5],
                "equipment": row[6],
            })

    wb.close()

    return {
        "service_mode": service_mode,
        "siso": siso_rows,
        "mimo": mimo_rows,
        "physical": physical,
    }


def find_threshold_row(
    rows: list[dict],
    bw_lte_mhz: int,
    bw_nr_c1_mhz: int,
    bw_nr_c2_mhz: int,
) -> dict | None:
    """Find the matching threshold row for a BW combination.

    Args:
        rows: List of row dicts from SISO or MIMO sheet
        bw_lte_mhz: LTE bandwidth in MHz
        bw_nr_c1_mhz: NR C1 bandwidth in MHz (0 if not present)
        bw_nr_c2_mhz: NR C2 bandwidth in MHz (0 if not present)

    Returns:
        Matching row dict, or None if no match found
    """
    for row in rows:
        if (row.get("bw_lte") == bw_lte_mhz
                and row.get("bw_nr_c1") == bw_nr_c1_mhz
                and row.get("bw_nr_c2") == bw_nr_c2_mhz):
            return row
    return None


def get_progressive_ul_threshold(row: dict, sinr_db: float) -> float | None:
    """Get the progressive UL threshold for a given SINR level.

    For EN-DC UL, the expected throughput varies by measured SINR.
    Uses the nearest lower SINR level from the 8 progressive levels.

    Args:
        row: Threshold row dict
        sinr_db: Measured SINR in dB

    Returns:
        UL threshold in Mbps, or None if SINR too low for progressive lookup
    """
    if sinr_db < PROGRESSIVE_UL_SINR_LEVELS[0]:
        return None

    # Find nearest lower SINR level
    selected_sinr = PROGRESSIVE_UL_SINR_LEVELS[0]
    for level in PROGRESSIVE_UL_SINR_LEVELS:
        if level <= sinr_db:
            selected_sinr = level
        else:
            break

    col_key = f"ul_sinr_{selected_sinr}"
    return row.get(col_key)
