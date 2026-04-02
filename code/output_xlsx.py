"""Output Excel generator — creates the 16-column Output.xlsx with threshold reference tab.

Tab 1: Validation Results
  BTS | Tech/Sector | Connection Mode | Bandwidth | PCI | RSRP | RSRQ | SINR |
  UE TX Power | SM-ST Duration | DL Throughput | UL Throughput | Comment (PASS/FAIL) |
  Observations | Recommendations | Impact on KPIs

Tab 2: Threshold Reference (copy of DAS_Validation_Thresholds.xlsx sheets)
  - SISO Speed Test (16 rows)
  - MIMO Speed Test (16 rows)
  - Service Mode Thresholds

Row count is DYNAMIC — depends on (sectors × technologies per sector).

Implementation: Claude Code Prompt 8 (Output Generation)
"""
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .utils.text_utils import truncate_for_cell

logger = logging.getLogger(__name__)

# 16 output columns (A–P)
COLUMNS = [
    "BTS",
    "Tech/Sector",
    "Connection Mode",
    "Bandwidth",
    "PCI",
    "RSRP",
    "RSRQ",
    "SINR",
    "UE TX Power",
    "SM-ST Duration",
    "DL Throughput",
    "UL Throughput",
    "Comment",
    "Observations",
    "Recommendations",
    "Impact on KPIs",
]

# Suggested column widths (characters)
COLUMN_WIDTHS = {
    "BTS": 14,
    "Tech/Sector": 18,
    "Connection Mode": 16,
    "Bandwidth": 28,
    "PCI": 8,
    "RSRP": 10,
    "RSRQ": 10,
    "SINR": 10,
    "UE TX Power": 13,
    "SM-ST Duration": 15,
    "DL Throughput": 15,
    "UL Throughput": 15,
    "Comment": 40,
    "Observations": 50,
    "Recommendations": 50,
    "Impact on KPIs": 50,
}

# Fast mode: first 13 columns (BTS through Comment); Full mode: all 16
FAST_COLUMNS = COLUMNS[:13]
FULL_COLUMNS = COLUMNS

# Wrap-text columns
WRAP_COLUMNS = {"Observations", "Recommendations", "Impact on KPIs", "Comment"}

# Style constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FAIL_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Amber — some pass, some fail
EXTRACTION_FAILED_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# RSRP color gradient (excellent → poor)
RSRP_FILLS = {
    "excellent": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
    "good": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "fair": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "poor": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}

# SINR color gradient
SINR_FILLS = {
    "excellent": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
    "good": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "fair": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "poor": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}


class OutputXlsxGenerator:
    """Creates the Output.xlsx workbook with validation results and threshold reference."""

    def __init__(self, config: dict):
        self.config = config
        self.output_cfg = config.get("output", {})
        self.site_name = self.output_cfg.get("site_name", "UNKNOWN")

    def generate(
        self,
        results: list[dict],
        threshold_data: dict,
        output_path: str | Path,
        failed: list[dict] | None = None,
        mode: str = "fast",
    ) -> Path:
        """Create the complete Output.xlsx workbook.

        Args:
            results: List of cell result dicts (one per tested cell).
            threshold_data: Dict from mop_thresholds.load_threshold_excel()
                            with keys: service_mode, siso, mimo, physical.
            output_path: File path for the output workbook.
            failed: List of extraction-failed result dicts (orange-shaded rows).
            mode: "fast" (13 cols) or "full" (16 cols).

        Returns:
            Path to the written file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self._active_columns = FULL_COLUMNS if mode == "full" else FAST_COLUMNS

        wb = Workbook()

        # Tab 1 — Validation Results (default sheet)
        ws_results = wb.active
        ws_results.title = "Validation Results"
        self._create_results_tab(ws_results, results, failed=failed or [])

        # Tab 2 — Thresholds Reference
        ws_thresh = wb.create_sheet("Thresholds Reference")
        self._create_threshold_tab(ws_thresh, threshold_data)

        # Tab 3 — Physical Layer (only if any result has physical_results)
        physical_rows = [r.get("physical_results") for r in results if r.get("physical_results")]
        if physical_rows:
            ws_phys = wb.create_sheet("Physical Layer")
            self._create_physical_layer_tab(ws_phys, results)

        wb.save(str(output_path))
        logger.info("Output XLSX written to %s (%d rows)", output_path, len(results))
        return output_path

    # ------------------------------------------------------------------
    # Tab 1: Validation Results
    # ------------------------------------------------------------------

    def _create_results_tab(
        self, ws: Worksheet, results: list[dict], failed: list[dict] | None = None,
    ) -> None:
        """Write Tab 1 with 16 columns, formatting, and conditional fills."""
        self._write_header_row(ws)

        for row_idx, cell_result in enumerate(results, start=2):
            self._write_data_row(ws, row_idx, cell_result)

        # Append orange-shaded rows for failed extractions
        next_row = len(results) + 2
        for fail_result in (failed or []):
            self._write_failed_row(ws, next_row, fail_result)
            next_row += 1

        self._set_column_widths(ws)

        # Enable auto-filter on header row
        last_col = get_column_letter(len(self._active_columns))
        total_rows = len(results) + len(failed or [])
        last_row = max(total_rows + 1, 2)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # Freeze top row
        ws.freeze_panes = "A2"

    def _write_header_row(self, ws: Worksheet) -> None:
        """Write bold, styled header row using active column set."""
        for col_idx, col_name in enumerate(self._active_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    def _write_data_row(self, ws: Worksheet, row_num: int, cell_result: dict) -> None:
        """Write one row of validation results and apply conditional formatting."""
        # Extract values from the result dict
        all_values = [
            cell_result.get("bts", self.site_name),
            cell_result.get("tech_sector", ""),
            cell_result.get("connection_mode", ""),
            cell_result.get("bandwidth", ""),
            cell_result.get("pci", ""),
            cell_result.get("rsrp"),
            cell_result.get("rsrq"),
            cell_result.get("sinr"),
            cell_result.get("tx_power"),
            cell_result.get("sm_st_duration"),
            cell_result.get("dl_throughput"),
            cell_result.get("ul_throughput"),
            cell_result.get("comment", ""),
            truncate_for_cell(cell_result.get("observations", "")),
            truncate_for_cell(cell_result.get("recommendations", "")),
            truncate_for_cell(cell_result.get("kpi_impact", "")),
        ]
        values = all_values[: len(self._active_columns)]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = THIN_BORDER
            col_name = self._active_columns[col_idx - 1]

            # Wrap text for long-content columns
            if col_name in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")

            # Number formatting for numeric columns
            if col_name in ("RSRP", "RSRQ", "SINR", "UE TX Power"):
                if isinstance(value, (int, float)):
                    cell.number_format = "0.0"
            elif col_name in ("DL Throughput", "UL Throughput"):
                if isinstance(value, (int, float)):
                    cell.number_format = "0.00"
            elif col_name == "SM-ST Duration":
                if isinstance(value, (int, float)):
                    cell.number_format = "0"

        # --- Conditional formatting ---

        # DL Throughput (K) / UL Throughput (L) — individual pass/fail coloring
        dl_pf = cell_result.get("dl_pass_fail")
        ul_pf = cell_result.get("ul_pass_fail")
        dl_col = self._active_columns.index("DL Throughput") + 1 if "DL Throughput" in self._active_columns else None
        ul_col = self._active_columns.index("UL Throughput") + 1 if "UL Throughput" in self._active_columns else None

        if dl_col and dl_pf == "PASS":
            ws.cell(row=row_num, column=dl_col).fill = PASS_FILL
        elif dl_col and dl_pf == "FAIL":
            ws.cell(row=row_num, column=dl_col).fill = FAIL_FILL

        if ul_col and ul_pf == "PASS":
            ws.cell(row=row_num, column=ul_col).fill = PASS_FILL
        elif ul_col and ul_pf == "FAIL":
            ws.cell(row=row_num, column=ul_col).fill = FAIL_FILL

        # Comment column — green/amber/red based on DL+UL results
        comment = str(cell_result.get("comment", "")).upper()
        comment_col = self._active_columns.index("Comment") + 1 if "Comment" in self._active_columns else 13
        comment_cell = ws.cell(row=row_num, column=comment_col)
        if "PASS" in comment and "FAIL" not in comment:
            comment_cell.fill = PASS_FILL
        elif "FAIL" in comment:
            # Partial fail (one passes, one fails) → amber; both fail → red
            dl_ok = dl_pf == "PASS"
            ul_ok = ul_pf == "PASS"
            if dl_ok or ul_ok:
                comment_cell.fill = PARTIAL_FAIL_FILL
            else:
                comment_cell.fill = FAIL_FILL

        # SM-ST Duration (J) — pass/fail (≤4 min = PASS, >4 min = FAIL)
        dur_pf = cell_result.get("duration_pass_fail")
        dur_col = self._active_columns.index("SM-ST Duration") + 1 if "SM-ST Duration" in self._active_columns else None
        if dur_col and dur_pf == "PASS":
            ws.cell(row=row_num, column=dur_col).fill = PASS_FILL
        elif dur_col and dur_pf == "FAIL":
            ws.cell(row=row_num, column=dur_col).fill = FAIL_FILL

        # RSRP (F) — quality gradient
        rsrp = cell_result.get("rsrp")
        if isinstance(rsrp, (int, float)):
            fill = self._get_rsrp_fill(rsrp)
            if fill:
                ws.cell(row=row_num, column=6).fill = fill

        # SINR (H) — quality gradient
        sinr = cell_result.get("sinr")
        if isinstance(sinr, (int, float)):
            fill = self._get_sinr_fill(sinr)
            if fill:
                ws.cell(row=row_num, column=8).fill = fill

        # RSRQ (G) — quality gradient
        rsrq = cell_result.get("rsrq")
        if isinstance(rsrq, (int, float)):
            fill = self._get_rsrq_fill(rsrq)
            if fill:
                ws.cell(row=row_num, column=7).fill = fill

    def _write_failed_row(self, ws: Worksheet, row_num: int, fail_result: dict) -> None:
        """Write an orange-shaded row for a failed extraction."""
        all_values = [
            fail_result.get("cell_id", ""),
            fail_result.get("tech_subfolder", ""),
            "",  # connection_mode
            "",  # bandwidth
            "",  # pci
            None, None, None, None, None, None, None,  # RF params + throughput
            f"EXTRACTION_FAILED: {fail_result.get('error', 'unknown')}",
            "", "", "",
        ]
        values = all_values[: len(self._active_columns)]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.fill = EXTRACTION_FAILED_FILL
            col_name = self._active_columns[col_idx - 1]
            if col_name in WRAP_COLUMNS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")

    # ------------------------------------------------------------------
    # Tab 2: Thresholds Reference
    # ------------------------------------------------------------------

    def _create_threshold_tab(self, ws: Worksheet, threshold_data: dict) -> None:
        """Write threshold reference tables: SISO, MIMO, Service Mode."""
        current_row = 1

        # --- SISO Speed Test ---
        current_row = self._write_speed_table(
            ws, current_row, "SISO Speed Test Thresholds", threshold_data.get("siso", [])
        )

        # Blank separator row
        current_row += 2

        # --- MIMO Speed Test ---
        current_row = self._write_speed_table(
            ws, current_row, "MIMO Speed Test Thresholds", threshold_data.get("mimo", [])
        )

        # Blank separator row
        current_row += 2

        # --- Service Mode Thresholds ---
        current_row = self._write_service_mode_table(
            ws, current_row, threshold_data.get("service_mode", {})
        )

        # Auto-size columns
        for col_idx in range(1, 30):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14

    def _write_speed_table(
        self, ws: Worksheet, start_row: int, title: str, rows: list[dict]
    ) -> int:
        """Write a SISO or MIMO speed test threshold table.

        Returns:
            Next available row number.
        """
        # Title row
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        start_row += 1

        # Header row
        speed_headers = [
            "Config", "BW LTE", "BW NR C1", "BW NR C2",
            "LTE DL Min", "LTE DL Max", "NR DL Min", "NR DL Max",
            "EN-DC DL Min", "EN-DC DL Max", "NR-DC DL Min", "NR-DC DL Max",
            "LTE UL Min", "LTE UL Max", "NR UL Min", "NR UL Max",
            "EN-DC UL Min", "EN-DC UL Max",
            "UL SINR 15", "UL SINR 19", "UL SINR 20", "UL SINR 22",
            "UL SINR 24", "UL SINR 26", "UL SINR 28", "UL SINR 30",
        ]
        for col_idx, header in enumerate(speed_headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.border = THIN_BORDER
        start_row += 1

        # Data rows
        col_keys = [
            "config", "bw_lte", "bw_nr_c1", "bw_nr_c2",
            "lte_dl_min", "lte_dl_max", "nr_dl_min", "nr_dl_max",
            "endc_dl_min", "endc_dl_max", "nrdc_dl_min", "nrdc_dl_max",
            "lte_ul_min", "lte_ul_max", "nr_ul_min", "nr_ul_max",
            "endc_ul_min", "endc_ul_max",
            "ul_sinr_15", "ul_sinr_19", "ul_sinr_20", "ul_sinr_22",
            "ul_sinr_24", "ul_sinr_26", "ul_sinr_28", "ul_sinr_30",
        ]
        for row_data in rows:
            for col_idx, key in enumerate(col_keys, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=row_data.get(key))
                cell.border = THIN_BORDER
                if isinstance(row_data.get(key), (int, float)):
                    cell.number_format = "0.0"
            start_row += 1

        return start_row

    def _write_service_mode_table(
        self, ws: Worksheet, start_row: int, service_mode: dict
    ) -> int:
        """Write Service Mode threshold reference table.

        Returns:
            Next available row number.
        """
        title_cell = ws.cell(row=start_row, column=1, value="Service Mode Thresholds")
        title_cell.font = Font(bold=True, size=12)
        start_row += 1

        # Header
        sm_headers = [
            "Location", "RSRP Min", "RSRP Max", "SINR Min",
            "RSRQ Min", "RSRQ Max", "TX Power",
        ]
        for col_idx, header in enumerate(sm_headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.border = THIN_BORDER
        start_row += 1

        # Data rows
        sm_keys = ["rsrp_min", "rsrp_max", "sinr_min", "rsrq_min", "rsrq_max", "tx_power"]
        for location, params in service_mode.items():
            ws.cell(row=start_row, column=1, value=location).border = THIN_BORDER
            for col_idx, key in enumerate(sm_keys, start=2):
                cell = ws.cell(row=start_row, column=col_idx, value=params.get(key))
                cell.border = THIN_BORDER
            start_row += 1

        return start_row

    def _set_column_widths(self, ws: Worksheet) -> None:
        """Set column widths based on COLUMN_WIDTHS mapping."""
        for col_idx, col_name in enumerate(self._active_columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 14)

    # ------------------------------------------------------------------
    # Tab 3: Physical Layer
    # ------------------------------------------------------------------

    def _create_physical_layer_tab(
        self, ws: Worksheet, results: list[dict],
    ) -> None:
        """Write Physical Layer check results for cells that have them."""
        headers = [
            "BTS", "Tech/Sector", "Parameter", "Value",
            "Min", "Max", "Unit", "Band", "Equipment",
            "Pass/Fail", "Delta",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        row_num = 2
        for result in results:
            phys = result.get("physical_results")
            if not phys:
                continue
            bts = result.get("bts", "")
            tech_sector = result.get("tech_sector", "")
            for pr in phys:
                values = [
                    bts,
                    tech_sector,
                    pr.get("parameter", ""),
                    pr.get("value"),
                    pr.get("min_value"),
                    pr.get("max_value"),
                    pr.get("unit", ""),
                    pr.get("band", ""),
                    pr.get("equipment", ""),
                    pr.get("pass_fail", ""),
                    pr.get("delta"),
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = THIN_BORDER
                    if isinstance(val, (int, float)):
                        cell.number_format = "0.00"

                # Color the Pass/Fail column
                pf_cell = ws.cell(row=row_num, column=10)
                if pr.get("pass_fail") == "PASS":
                    pf_cell.fill = PASS_FILL
                elif pr.get("pass_fail") == "FAIL":
                    pf_cell.fill = FAIL_FILL

                row_num += 1

        # Column widths
        widths = [14, 18, 18, 10, 10, 10, 8, 14, 16, 10, 10]
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # Color helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _get_rsrp_fill(value: float) -> PatternFill | None:
        if value >= -40:
            return RSRP_FILLS["excellent"]
        elif value >= -75:
            return RSRP_FILLS["good"]
        elif value >= -90:
            return RSRP_FILLS["fair"]
        elif value >= -140:
            return RSRP_FILLS["poor"]
        return None

    @staticmethod
    def _get_sinr_fill(value: float) -> PatternFill | None:
        if value >= 25:
            return SINR_FILLS["excellent"]
        elif value >= 15:
            return SINR_FILLS["good"]
        elif value >= 5:
            return SINR_FILLS["fair"]
        elif value >= -20:
            return SINR_FILLS["poor"]
        return None

    @staticmethod
    def _get_rsrq_fill(value: float) -> PatternFill | None:
        if value >= -3:
            return RSRP_FILLS["excellent"]
        elif value >= -7:
            return RSRP_FILLS["good"]
        elif value >= -12:
            return RSRP_FILLS["fair"]
        elif value >= -20:
            return RSRP_FILLS["poor"]
        return None
