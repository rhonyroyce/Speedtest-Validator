"""Excel utilities — openpyxl helper functions for reading and writing.

Shared utilities for CIQ reading, threshold loading, and output generation.

Implementation: Claude Code Prompt 4 (CIQ Reader) — shared utility
"""
import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


def read_sheet_as_dicts(
    workbook_path: str | Path,
    sheet_name: str,
    header_row: int = 1,
) -> list[dict]:
    """Read an Excel sheet into a list of dicts (one per row).

    Args:
        workbook_path: Path to .xlsx file
        sheet_name: Sheet name to read
        header_row: Row number containing column headers (1-indexed)

    Returns:
        List of dicts keyed by header names
    """
    wb = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path}. "
                             f"Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < header_row:
        return []

    headers = rows[header_row - 1]
    # Strip None headers, keep track of valid column indices
    col_map = [(i, str(h).strip()) for i, h in enumerate(headers) if h is not None]

    results = []
    for row in rows[header_row:]:
        # Skip entirely empty rows
        if not any(v is not None for v in row):
            continue
        d = {}
        for col_idx, col_name in col_map:
            d[col_name] = row[col_idx] if col_idx < len(row) else None
        results.append(d)

    logger.debug("Read %d rows from sheet '%s' in %s", len(results), sheet_name, workbook_path)
    return results


def find_column_index(sheet, column_name: str, header_row: int = 1) -> int | None:
    """Find the column index (0-based) for a given header name.

    Case-insensitive, strips whitespace.

    Args:
        sheet: openpyxl worksheet
        column_name: Column header to find
        header_row: Row number containing headers (1-indexed)

    Returns:
        0-based column index, or None if not found
    """
    target = column_name.strip().lower()
    for row in sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for i, cell_val in enumerate(row):
            if cell_val is not None and str(cell_val).strip().lower() == target:
                return i
    return None


def convert_bw_khz_to_mhz(bw_khz: int | float) -> float:
    """Convert bandwidth from kHz (CIQ format) to MHz (display format).

    CIQ stores bandwidth in kHz: 20000 = 20 MHz, 100000 = 100 MHz.

    Args:
        bw_khz: Bandwidth in kHz

    Returns:
        Bandwidth in MHz
    """
    return round(bw_khz / 1000, 1)


def safe_float(value, default: float = 0.0) -> float:
    """Safely convert a cell value to float, returning default on failure."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_int(value, default: int = 0) -> int:
    """Safely convert a cell value to int, returning default on failure."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default
