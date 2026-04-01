"""Tests for output_xlsx.py — verify generator creates valid files.

Implementation: Claude Code Prompt 8 (Output Generation)
"""
import sys
from pathlib import Path

import pytest
import yaml

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from code.output_xlsx import OutputXlsxGenerator

CONFIG_PATH = Path(__file__).resolve().parent.parent / "config.yaml"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / "test"


@pytest.fixture
def config():
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


@pytest.fixture
def mock_results():
    """Minimal mock cell results — 3 cells covering PASS, FAIL, and edge cases."""
    return [
        {
            "bts": "SFY0803A",
            "tech_sector": "L19/Sector 1",
            "connection_mode": "LTE Only",
            "bandwidth": "20 MHz LTE",
            "pci": 101,
            "rsrp": -55.0,
            "rsrq": -5.0,
            "sinr": 28.0,
            "tx_power": -10.0,
            "sm_st_duration": 15,
            "dl_throughput": 85.5,
            "ul_throughput": 25.3,
            "comment": "PASS — DL 85.5 Mbps (min 50), UL 25.3 Mbps (min 15)",
            "observations": "RSRP -55 dBm — good signal within DAS thresholds. SINR 28 dB — excellent.",
            "recommendations": "No action required. All parameters within expected range.",
            "kpi_impact": "No negative KPI impact expected. Capacity and accessibility nominal.",
        },
        {
            "bts": "SFY0803A",
            "tech_sector": "N2500_C1 NSA/Sector 2",
            "connection_mode": "EN-DC",
            "bandwidth": "20 MHz LTE + 100 MHz NR",
            "pci": 202,
            "rsrp": -82.0,
            "rsrq": -9.5,
            "sinr": 12.0,
            "tx_power": 5.0,
            "sm_st_duration": 30,
            "dl_throughput": 120.0,
            "ul_throughput": 8.5,
            "comment": "FAIL — UL 8.5 Mbps (min 20). Delta: -11.5 Mbps",
            "observations": "RSRP -82 dBm — fair signal. SINR 12 dB — moderate interference. TX Power +5 dBm — ELEVATED.",
            "recommendations": "Investigate UL path loss. Check antenna feeder and splitter connections. TX power compensation indicates coverage issue.",
            "kpi_impact": "Potential ACC and RET KPI degradation. High TX power impacts PWR domain.",
        },
        {
            "bts": "SFY0803A",
            "tech_sector": "N2500_C1 SA/Sector 3",
            "connection_mode": "NR SA",
            "bandwidth": "100 MHz NR",
            "pci": 303,
            "rsrp": -68.0,
            "rsrq": -6.0,
            "sinr": 30.0,
            "tx_power": -15.0,
            "sm_st_duration": 10,
            "dl_throughput": 450.0,
            "ul_throughput": 95.0,
            "comment": "PASS — DL 450 Mbps (min 300), UL 95 Mbps (min 50)",
            "observations": "RSRP -68 dBm — good signal. SINR 30 dB — excellent quality.",
            "recommendations": "No issues detected. Performance within expected range for NR SA 100 MHz.",
            "kpi_impact": "All KPI domains nominal. NR SA throughput meets capacity targets.",
        },
    ]


@pytest.fixture
def mock_threshold_data():
    """Mock threshold data matching mop_thresholds.load_threshold_excel() output."""
    return {
        "service_mode": {
            "Radio End": {
                "rsrp_min": -60, "rsrp_max": -40, "sinr_min": 25,
                "rsrq_min": -12, "rsrq_max": -3, "tx_power": "negative",
            },
            "Antenna End": {
                "rsrp_min": -75, "rsrp_max": -50, "sinr_min": 25,
                "rsrq_min": -12, "rsrq_max": -3, "tx_power": "negative",
            },
        },
        "siso": [
            {
                "config": "SISO", "bw_lte": 20, "bw_nr_c1": 0, "bw_nr_c2": 0,
                "lte_dl_min": 50, "lte_dl_max": 75, "nr_dl_min": None, "nr_dl_max": None,
                "endc_dl_min": None, "endc_dl_max": None, "nrdc_dl_min": None, "nrdc_dl_max": None,
                "lte_ul_min": 15, "lte_ul_max": 30, "nr_ul_min": None, "nr_ul_max": None,
                "endc_ul_min": None, "endc_ul_max": None,
                "ul_sinr_15": None, "ul_sinr_19": None, "ul_sinr_20": None, "ul_sinr_22": None,
                "ul_sinr_24": None, "ul_sinr_26": None, "ul_sinr_28": None, "ul_sinr_30": None,
            },
        ],
        "mimo": [
            {
                "config": "MIMO", "bw_lte": 20, "bw_nr_c1": 100, "bw_nr_c2": 0,
                "lte_dl_min": 75, "lte_dl_max": 150, "nr_dl_min": 300, "nr_dl_max": 600,
                "endc_dl_min": 350, "endc_dl_max": 700, "nrdc_dl_min": None, "nrdc_dl_max": None,
                "lte_ul_min": 20, "lte_ul_max": 40, "nr_ul_min": 50, "nr_ul_max": 120,
                "endc_ul_min": 20, "endc_ul_max": 50,
                "ul_sinr_15": 10, "ul_sinr_19": 15, "ul_sinr_20": 18, "ul_sinr_22": 22,
                "ul_sinr_24": 28, "ul_sinr_26": 35, "ul_sinr_28": 42, "ul_sinr_30": 50,
            },
        ],
        "physical": [],
    }


@pytest.fixture
def mock_site_config():
    """Mock CIQ site configuration rows."""
    return [
        {"sector": "1", "technology": "LTE", "band": "B19", "bandwidth": "20 MHz", "mimo_config": "MIMO", "pci": 101, "earfcn": 6300},
        {"sector": "2", "technology": "NR", "band": "n41", "bandwidth": "100 MHz", "mimo_config": "MIMO", "pci": 202, "arfcn": 520000},
        {"sector": "3", "technology": "NR", "band": "n41", "bandwidth": "100 MHz", "mimo_config": "SISO", "pci": 303, "arfcn": 520000},
    ]


@pytest.fixture
def mock_analysis_data():
    """Mock aggregated analysis data for docx generator."""
    return {
        "site_id": "SFY0803A",
        "date": "2026-03-19",
        "kpi_impacts": [
            {
                "cell": "N2500_C1 NSA/Sector 2",
                "kpi_impact": "High TX power (+5 dBm) indicates UL coverage gap. ACC degradation from RRC setup failures. RET risk from radio link failures. PWR impact from UE battery drain.",
            },
        ],
    }


# =====================================================================
# OutputXlsxGenerator Tests
# =====================================================================

class TestOutputXlsxGenerate:
    """Test that generate() creates a valid .xlsx file."""

    def test_creates_file(self, config, mock_results, mock_threshold_data, tmp_path):
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        result = gen.generate(mock_results, mock_threshold_data, out)
        assert result.exists()
        assert result.stat().st_size > 0

    def test_two_sheets(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out)
        wb = load_workbook(str(out))
        assert "Validation Results" in wb.sheetnames
        assert "Thresholds Reference" in wb.sheetnames
        wb.close()

    def test_correct_row_count(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out)
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        # Header row + 3 data rows
        assert ws.max_row == 4
        wb.close()

    def test_header_values(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out, mode="full")
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 17)]
        from code.output_xlsx import COLUMNS
        assert headers == COLUMNS
        wb.close()

    def test_pass_fill_green(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out)
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        # Row 2 (first data row) should be PASS → green fill on Comment (col M=13)
        comment_cell = ws.cell(row=2, column=13)
        assert comment_cell.fill.start_color.rgb == "00C6EFCE"
        wb.close()

    def test_fail_fill_red(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out)
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        # Row 3 (second data row) should be FAIL → red fill on Comment
        comment_cell = ws.cell(row=3, column=13)
        assert comment_cell.fill.start_color.rgb == "00FFC7CE"
        wb.close()

    def test_threshold_tab_has_data(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "test_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out)
        wb = load_workbook(str(out))
        ws = wb["Thresholds Reference"]
        # Should have content (SISO title + header + data + MIMO title + header + data + SM title + header + data)
        assert ws.max_row > 5
        wb.close()

    def test_empty_results(self, config, mock_threshold_data, tmp_path):
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "empty_output.xlsx"
        result = gen.generate([], mock_threshold_data, out)
        assert result.exists()


# =====================================================================
# Cell Data Flattening Tests
# =====================================================================

class TestCellDataFlattening:
    """Verify Phase 5 cell_data flattening logic (unit-level, no LLM)."""

    def test_cell_data_flattening_lte(self):
        """LTE-only result produces flat cell_data with correct fields."""
        result = {
            "service_mode": {
                "lte_params": {"rsrp_dbm": -55.0, "sinr_db": 28.0, "rsrq_db": -5.0, "tx_power_dbm": -10.0},
                "nr_params": {},
            },
            "speedtest": {"dl_throughput_mbps": 85.5, "ul_throughput_mbps": 25.3},
            "inferred_conn_mode": "LTE Only",
            "bw_lte_mhz": 20,
            "bw_nr_c1_mhz": 0,
            "bw_nr_c2_mhz": 0,
            "mimo_config": "MIMO",
        }
        sm = result.get("service_mode") or {}
        st = result.get("speedtest") or {}
        lte = sm.get("lte_params") or {}
        nr = sm.get("nr_params") or {}
        cell_data = {
            "rsrp": lte.get("rsrp_dbm") or nr.get("nr5g_rsrp_dbm"),
            "sinr": lte.get("sinr_db") or nr.get("nr5g_sinr_db"),
            "rsrq": lte.get("rsrq_db") or nr.get("nr5g_rsrq_db"),
            "tx_power": lte.get("tx_power_dbm") or nr.get("nr_tx_power_dbm"),
            "dl_throughput": st.get("dl_throughput_mbps"),
            "ul_throughput": st.get("ul_throughput_mbps"),
            "tech": "NR" if nr.get("nr_band") else "LTE",
            "conn_mode": result.get("inferred_conn_mode", "LTE Only"),
            "bw_lte_mhz": result.get("bw_lte_mhz", 0),
            "bw_nr_c1_mhz": result.get("bw_nr_c1_mhz", 0),
            "bw_nr_c2_mhz": result.get("bw_nr_c2_mhz", 0),
        }
        assert cell_data["rsrp"] == -55.0
        assert cell_data["sinr"] == 28.0
        assert cell_data["dl_throughput"] == 85.5
        assert cell_data["tech"] == "LTE"
        assert cell_data["conn_mode"] == "LTE Only"

    def test_cell_data_flattening_nr(self):
        """NR result with nr_band populates NR fields and tech='NR'."""
        result = {
            "service_mode": {
                "lte_params": {},
                "nr_params": {
                    "nr5g_rsrp_dbm": -68.0,
                    "nr5g_sinr_db": 30.0,
                    "nr5g_rsrq_db": -6.0,
                    "nr_tx_power_dbm": -15.0,
                    "nr_band": "n41",
                },
            },
            "speedtest": {"dl_throughput_mbps": 450.0, "ul_throughput_mbps": 95.0},
            "inferred_conn_mode": "NR SA",
            "bw_lte_mhz": 0,
            "bw_nr_c1_mhz": 100,
            "bw_nr_c2_mhz": 0,
        }
        sm = result.get("service_mode") or {}
        st = result.get("speedtest") or {}
        lte = sm.get("lte_params") or {}
        nr = sm.get("nr_params") or {}
        cell_data = {
            "rsrp": lte.get("rsrp_dbm") or nr.get("nr5g_rsrp_dbm"),
            "sinr": lte.get("sinr_db") or nr.get("nr5g_sinr_db"),
            "rsrq": lte.get("rsrq_db") or nr.get("nr5g_rsrq_db"),
            "tx_power": lte.get("tx_power_dbm") or nr.get("nr_tx_power_dbm"),
            "dl_throughput": st.get("dl_throughput_mbps"),
            "ul_throughput": st.get("ul_throughput_mbps"),
            "tech": "NR" if nr.get("nr_band") else "LTE",
            "conn_mode": result.get("inferred_conn_mode", "LTE Only"),
        }
        assert cell_data["rsrp"] == -68.0
        assert cell_data["sinr"] == 30.0
        assert cell_data["tech"] == "NR"
        assert cell_data["conn_mode"] == "NR SA"
        assert cell_data["dl_throughput"] == 450.0


# =====================================================================
# Fast / Full Mode Column Tests
# =====================================================================

class TestFastFullModeColumns:
    """Verify fast mode produces 13 columns and full mode produces 16."""

    def test_fast_mode_13_cols(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        from code.output_xlsx import FAST_COLUMNS
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "fast_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out, mode="fast")
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 14)]
        assert headers == list(FAST_COLUMNS)
        assert len(headers) == 13
        # Column 14 should be empty (no LLM columns)
        assert ws.cell(row=1, column=14).value is None
        wb.close()

    def test_full_mode_16_cols(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        from code.output_xlsx import FULL_COLUMNS
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "full_output.xlsx"
        gen.generate(mock_results, mock_threshold_data, out, mode="full")
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, 17)]
        assert headers == list(FULL_COLUMNS)
        assert len(headers) == 16
        wb.close()

    def test_fast_mode_data_row_13_values(self, config, mock_results, mock_threshold_data, tmp_path):
        from openpyxl import load_workbook
        gen = OutputXlsxGenerator(config)
        out = tmp_path / "fast_data.xlsx"
        gen.generate(mock_results, mock_threshold_data, out, mode="fast")
        wb = load_workbook(str(out))
        ws = wb["Validation Results"]
        # Row 2 (first data row) should have 13 non-None header-aligned values
        row_vals = [ws.cell(row=2, column=i).value for i in range(1, 14)]
        assert row_vals[0] is not None  # BTS
        # Column 14 should be empty
        assert ws.cell(row=2, column=14).value is None
        wb.close()
