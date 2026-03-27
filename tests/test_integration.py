"""Integration tests — end-to-end pipeline with mocked Ollama.

Tests the full DASValidator pipeline (Phases 0–6) using fixture JSON
responses instead of real VLM/LLM calls.
"""
import json
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

FIXTURES = Path(__file__).parent / "fixtures"


def _load_fixture(name: str) -> dict:
    return json.loads((FIXTURES / name).read_text())


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_ollama(monkeypatch):
    """Patch OllamaClient so no real HTTP calls are made.

    - health_check → True
    - validate_models_available → [] (no missing models)
    - ensure_model_loaded / unload_model / get_loaded_models → no-op
    - chat_vision_json → returns fixture based on prompt content
    - chat_text → returns fixture analysis text
    """
    from code.ollama_client import OllamaClient

    sm_lte = _load_fixture("vlm_response_sm_lte.json")
    sm_nr = _load_fixture("vlm_response_sm_nr.json")
    st = _load_fixture("vlm_response_st.json")
    analysis = _load_fixture("analysis_response.json")

    monkeypatch.setattr(OllamaClient, "health_check", lambda self: True)
    monkeypatch.setattr(OllamaClient, "validate_models_available", lambda self: [])
    monkeypatch.setattr(OllamaClient, "ensure_model_loaded", lambda self, m: None)
    monkeypatch.setattr(OllamaClient, "unload_model", lambda self, m: None)
    monkeypatch.setattr(OllamaClient, "get_loaded_models", lambda self: [])

    def fake_chat_vision_json(self, image_b64, prompt, model=None, temperature=None):
        # Decide which fixture to return based on prompt content
        if "speedtest" in prompt.lower():
            return st, 1
        elif "service" in prompt.lower():
            # Alternate between LTE and NR based on a counter
            if not hasattr(fake_chat_vision_json, "_call_count"):
                fake_chat_vision_json._call_count = 0
            fake_chat_vision_json._call_count += 1
            if fake_chat_vision_json._call_count % 2 == 0:
                return sm_nr, 1
            return sm_lte, 1
        return sm_lte, 1

    monkeypatch.setattr(OllamaClient, "chat_vision_json", fake_chat_vision_json)

    def fake_chat_text(self, prompt, model=None, temperature=None):
        if "observation" in prompt.lower():
            return analysis["observations"]
        elif "recommendation" in prompt.lower():
            return analysis["recommendations"]
        elif "kpi" in prompt.lower():
            return analysis["kpi_impact"]
        return analysis["observations"]

    monkeypatch.setattr(OllamaClient, "chat_text", fake_chat_text)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_full_pipeline_dry_run(mock_ollama, tmp_path):
    """Run the full pipeline in dry-run mode and verify output xlsx is created."""
    from code.main import DASValidator

    validator = DASValidator(config_path="config.yaml")
    validator.run(
        site_folder="./input/SFY0803A",
        ciq_path="./input/SFY0803A_MMBB_CIQ_EXPORT_20251127_173752.xlsx",
        output_dir=str(tmp_path),
        dry_run=True,
        mode="fast",
    )

    output = tmp_path / "SFY0803A_Output.xlsx"
    assert output.exists(), "Output xlsx was not created"

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    assert ws.max_row >= 3, f"Expected header + >=2 data rows, got {ws.max_row} rows"


def test_full_pipeline_full_mode(mock_ollama, tmp_path):
    """Run the pipeline in full mode (with Phase 5 analysis) in dry-run."""
    from code.main import DASValidator

    validator = DASValidator(config_path="config.yaml")
    validator.run(
        site_folder="./input/SFY0803A",
        ciq_path="./input/SFY0803A_MMBB_CIQ_EXPORT_20251127_173752.xlsx",
        output_dir=str(tmp_path),
        dry_run=True,
        mode="full",
    )

    output = tmp_path / "SFY0803A_Output.xlsx"
    assert output.exists()

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Full mode should have 16 columns
    assert ws.max_column == 16, f"Expected 16 columns in full mode, got {ws.max_column}"


def test_connection_mode_propagation(mock_ollama, tmp_path):
    """Verify connection modes detected in Phase 2 propagate to Phase 6 output."""
    from code.main import DASValidator

    validator = DASValidator(config_path="config.yaml")
    validator.run(
        site_folder="./input/SFY0803A",
        ciq_path="./input/SFY0803A_MMBB_CIQ_EXPORT_20251127_173752.xlsx",
        output_dir=str(tmp_path),
        dry_run=True,
        mode="fast",
    )

    output = tmp_path / "SFY0803A_Output.xlsx"
    wb = openpyxl.load_workbook(output)
    ws = wb.active

    # Find Connection Mode column index (1-based)
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Connection Mode" in headers, f"Missing 'Connection Mode' column. Headers: {headers}"
    cm_col = headers.index("Connection Mode") + 1

    # Every data row should have a non-empty connection mode
    valid_modes = {"LTE Only", "NR SA", "EN-DC", "NR-DC"}
    for row in range(2, ws.max_row + 1):
        cm_value = ws.cell(row=row, column=cm_col).value
        assert cm_value is not None and cm_value != "", (
            f"Row {row} has empty Connection Mode"
        )
        assert cm_value in valid_modes, (
            f"Row {row} has unexpected Connection Mode: '{cm_value}'"
        )


def test_output_column_names(mock_ollama, tmp_path):
    """Verify all 16 expected columns exist in full-mode output."""
    from code.main import DASValidator

    expected_columns = [
        "BTS",
        "Tech/Sector",
        "Connection Mode",
        "Bandwidth",
        "PCI",
        "RSRP",
        "RSRQ",
        "SINR",
        "UE TX Power",
        "SM-ST Duration",
        "DL Throughput",
        "UL Throughput",
        "Comment",
        "Observations",
        "Recommendations",
        "Impact on KPIs",
    ]

    validator = DASValidator(config_path="config.yaml")
    validator.run(
        site_folder="./input/SFY0803A",
        ciq_path="./input/SFY0803A_MMBB_CIQ_EXPORT_20251127_173752.xlsx",
        output_dir=str(tmp_path),
        dry_run=True,
        mode="full",
    )

    output = tmp_path / "SFY0803A_Output.xlsx"
    wb = openpyxl.load_workbook(output)
    ws = wb.active

    actual_columns = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert actual_columns == expected_columns, (
        f"Column mismatch.\nExpected: {expected_columns}\nActual:   {actual_columns}"
    )
